import openpyxl
from pulp import *
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os

def process_food_data(file_path_DB, file_path_diet, sheet_name_1, sheet_name_2):
    sheet_1 = pd.read_excel(file_path_DB, sheet_name=sheet_name_1)
    sheet_2 = pd.read_excel(file_path_diet, sheet_name=sheet_name_2)

    sheet_1_to_dict = sheet_1.to_dict(orient='records')

    daily_intake_list = sheet_2["Unnamed: 1"].dropna().tolist()
    min_q = sheet_2["Unnamed: 2"].dropna().tolist()[1:]
    max_q = sheet_2["Unnamed: 3"].dropna().tolist()[1:]

    DB_alimenti = {}
    for elemento in sheet_1_to_dict:
        alimento = elemento['Alimento']
        valori = {
            "proteine": elemento['Proteine'],
            "carboidrati": elemento['Carboidrati'],
            "grassi": elemento['Grassi'],
            "calorie": elemento['Calorie']
        }
        DB_alimenti[alimento] = valori

    dizionario_alimenti_min = {}
    dizionario_alimenti_max = {}

    for alimento, minimo, massimo in zip(daily_intake_list, min_q, max_q):
        if alimento not in dizionario_alimenti_min and alimento not in dizionario_alimenti_max:
            dizionario_alimenti_min[alimento] = minimo / 100
            dizionario_alimenti_max[alimento] = massimo / 100
        else:
            dizionario_alimenti_min[alimento] += minimo / 100
            dizionario_alimenti_max[alimento] += massimo / 100

    final_daily_dict = {}
    for alimento in daily_intake_list:
        final_daily_dict[alimento] = DB_alimenti[alimento]

    sorted_alimenti_min = dict(sorted(dizionario_alimenti_min.items()))
    sorted_alimenti_max = dict(sorted(dizionario_alimenti_max.items()))

    return sorted_alimenti_min, sorted_alimenti_max, final_daily_dict, dizionario_alimenti_min, dizionario_alimenti_max


def pianificazione_pasti(alimenti, P_min, P_max, F_min, F_max, C_min, C_max, cal_min, cal_max, peso_min_alimenti, peso_max_alimenti, sorted_alimenti_min, sorted_alimenti_max):
    problema = LpProblem("PianificazionePasti", LpMaximize)

    alimenti_vars = LpVariable.dicts("", alimenti, lowBound=0, cat='Continuous')

    proteine_tot = lpSum([alimenti[i]['proteine'] * alimenti_vars[i] for i in alimenti])
    grassi_tot = lpSum([alimenti[i]['grassi'] * alimenti_vars[i] for i in alimenti])
    carboidrati_tot = lpSum([alimenti[i]['carboidrati'] * alimenti_vars[i] for i in alimenti])
    calorie_tot = lpSum([alimenti[i]['calorie'] * alimenti_vars[i] for i in alimenti])

    problema += proteine_tot  # La funzione obiettivo può essere modificata in base alle esigenze

    problema += proteine_tot >= P_min, "Proteine_min"
    problema += proteine_tot <= P_max, "Proteine_max"
    problema += grassi_tot >= F_min, "Grassi_min"
    problema += grassi_tot <= F_max, "Grassi_max"
    problema += carboidrati_tot >= C_min, "Carboidrati_min"
    problema += carboidrati_tot <= C_max, "Carboidrati_max"
    problema += calorie_tot >= cal_min, "Calorie_min"
    problema += calorie_tot <= cal_max, "Calorie_max"

    for alimento in alimenti:
        problema += alimenti_vars[alimento] >= peso_min_alimenti[alimento], f"Peso_min_{alimento}"
        problema += alimenti_vars[alimento] <= peso_max_alimenti[alimento], f"Peso_max_{alimento}"

    problema.solve()

    nomi_variabili = []
    valori_variabili = []
    for var in problema.variables():
        nomi_variabili.append(var.name.lstrip("_").replace("_", " "))
        valori_variabili.append(round(var.varValue * 100, 1))  # Moltiplica per 100 per ottenere il valore in grammi

    lista_combinata_ordinata = sorted(zip(nomi_variabili, valori_variabili))

    lista1_ordinata, lista2_ordinata = zip(*lista_combinata_ordinata)

    df_al = pd.DataFrame({
        'Alimento': lista1_ordinata,
        'Valore (g)': lista2_ordinata,
        'Range_min': list(np.array(list(sorted_alimenti_min.values()))*100),
        'Range_max': list(np.array(list(sorted_alimenti_max.values()))*100)
    })

    return df_al, proteine_tot, grassi_tot, carboidrati_tot, calorie_tot, problema


def create_excel(day, df_al, cal, proteine_tot, grassi_tot, carboidrati_tot, calorie_tot, file_path):
    file_path = file_path + str(cal) + ".xlsx"

    # Days of the week
    days_of_week = ["Lunedi", "Martedi", "Mercoledi", "Giovedi", "Venerdi", "Sabato", "Domenica"]

    if not os.path.exists(file_path):
        # Create a new workbook with a sheet for each day of the week
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for day_name in days_of_week:
                # Create an empty DataFrame for each day
                pd.DataFrame().to_excel(writer, sheet_name=day_name)

    # Load the existing workbook and set if_sheet_exists to 'replace'
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # Dropping unnecessary columns
        try:
            df_al.drop(["Range_min", "Range_max", " Violed"], axis=1, inplace=True)
        except KeyError as e:
            print(f"Errore durante la rimozione delle colonne: {e}")

        #df_al.to_csv(r'/Users/liviobasile/Downloads/Diet_builder/dataframe.csv', index=False)

        df_al.to_excel(writer, sheet_name=day, index=False)

        workbook = writer.book
        worksheet = workbook[day]

        num_rows = df_al.shape[0]

        # Add two empty rows after the data
        #worksheet.insert_rows(num_rows + 1, amount=1)

        worksheet.cell(row=num_rows + 3, column=1).value = "Apporto proteico totale:"
        worksheet.cell(row=num_rows + 3, column=2).value = f"{round(value(proteine_tot), 1)} g"

        worksheet.cell(row=num_rows + 4, column=1).value = "Apporto di grassi totale:"
        worksheet.cell(row=num_rows + 4, column=2).value = f"{round(value(grassi_tot), 1)} g"

        worksheet.cell(row=num_rows + 5, column=1).value = "Apporto di carboidrati totale:"
        worksheet.cell(row=num_rows + 5, column=2).value = f"{round(value(carboidrati_tot), 1)} g"

        worksheet.cell(row=num_rows + 6, column=1).value = "Apporto di calorie totale:"
        worksheet.cell(row=num_rows + 6, column=2).value = f"{round(value(calorie_tot), 1)} kcal"

    workbook.save(file_path)

def check_range(row):
    return '<-' if not (row['Range_min'] <= row['Valore (g)'] <= row['Range_max']) else ''
